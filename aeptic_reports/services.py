# aeptic_reports/services.py

from datetime import datetime, timedelta
from io import BytesIO
from decimal import Decimal

from django.utils import timezone as tz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from timetracking.models import TimeEntries, TimeEntryEvent
from requests.models import LeaveRequest
from admin.models import CompanySettings

# Importar reportlab para generación de PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# --- Mapeo de leave_reason a etiqueta legible ---
LEAVE_REASON_LABELS = {
    'sick': 'Baja por enfermedad',
    'maternity': 'Maternidad / Paternidad',
    'wedding': 'Matrimonio',
    'bereavement': 'Fallecimiento familiar',
    'medical_appointment': 'Cita médica',
    'legal_duty': 'Deber público / legal',
    'personal': 'Asuntos propios',
    'other': 'Otro',
}

# --- Columnas del Excel (orden A-K) ---
COLUMNS = [
    'Fecha',
    'Día Semana',
    'Hora Entrada',
    'Hora Salida',
    'Ausencia',
    'Vacaciones',
    'Baja',
    'Festivo',
    'Total Horas Ordinarias',
    'Total Horas Extras',
    'Firma Trabajador'
]

DAYS_OF_WEEK_ES = {
    0: 'Domingo',
    1: 'Lunes',
    2: 'Martes',
    3: 'Miércoles',
    4: 'Jueves',
    5: 'Viernes',
    6: 'Sábado',
}


# --- Helper Functions ---

def get_pause_hours_for_day(user_id, company_id, date):
    """Calcular total pausas del día en horas (suma de pause_start/pause_end)"""
    try:
        # Obtener TimeEntry para ese día
        time_entry = TimeEntries.objects.filter(
            user_id=user_id,
            company_id=company_id,
            date=date
        ).first()

        if not time_entry:
            return 0.0

        # Obtener todos los eventos de pausa para ese TimeEntry
        pause_events = TimeEntryEvent.objects.filter(
            time_entry=time_entry,
            event_type__in=['pause_start', 'pause_end']
        ).order_by('timestamp')

        if not pause_events.exists():
            return 0.0

        # Emparejar pause_start con pause_end y sumar duración
        total_pause_seconds = 0
        pause_start = None

        for event in pause_events:
            if event.event_type == 'pause_start':
                pause_start = event.timestamp
            elif event.event_type == 'pause_end' and pause_start:
                duration = event.timestamp - pause_start
                total_pause_seconds += duration.total_seconds()
                pause_start = None

        # Convertir a horas decimales (ej: 1.5 = 1 hora 30 min)
        return round(total_pause_seconds / 3600, 2)

    except Exception:
        return 0.0


def get_vacation_hours_for_day(user_id, company_id, date):
    """Obtener horas de vacaciones para un día específico

    Utiliza el multiplicador de horas (hour_multiplier) si está disponible
    para periodos vacacionales especiales.
    """
    try:
        leave_request = LeaveRequest.objects.filter(
            user_id=user_id,
            company_id=company_id,
            leave_type='vacation',
            status='approved',
            start_date__lte=date,
            end_date__gte=date
        ).first()

        if leave_request:
            # Obtener jornada laboral diaria
            daily_hours = get_work_hours_per_day(company_id)

            # Aplicar multiplicador (default 1.0)
            multiplier = leave_request.hour_multiplier or 1.0

            # Calcular horas finales
            vacation_hours = daily_hours * multiplier

            return round(vacation_hours, 2)

        return 0.0

    except Exception:
        return 0.0


def get_sick_leave_for_day(user_id, company_id, date):
    """Obtener baja para un día: retorna motivo (4 tipos específicos) o vacío.

    Solo devuelve: sick, maternity, wedding, bereavement
    Formato: "Baja por enfermedad" o "Maternidad / Paternidad", etc.
    """
    try:
        sick_leave_reasons = {'sick', 'maternity', 'wedding', 'bereavement'}

        leave_request = LeaveRequest.objects.filter(
            user_id=user_id,
            company_id=company_id,
            leave_type='absence',
            status='approved',
            start_date__lte=date,
            end_date__gte=date
        ).first()

        if leave_request and leave_request.leave_reason in sick_leave_reasons:
            return LEAVE_REASON_LABELS.get(leave_request.leave_reason, "")

        return ""

    except Exception:
        return ""


def get_time_off_for_day(user_id, company_id, date):
    """Obtener tiempo libre (ausencia) para un día: retorna (motivo, horas) o None.

    Solo devuelve: medical_appointment, legal_duty
    Formato: ("Cita médica", 2.5) o ("Deber público / legal", 4.0)

    Calcula las horas reales si start_time/end_time están definidas,
    sino devuelve la jornada laboral completa.
    """
    try:
        time_off_reasons = {'medical_appointment', 'legal_duty'}
        work_hours_per_day = get_work_hours_per_day(company_id)

        leave_request = LeaveRequest.objects.filter(
            user_id=user_id,
            company_id=company_id,
            leave_type='absence',
            status='approved',
            start_date__lte=date,
            end_date__gte=date
        ).first()

        if leave_request and leave_request.leave_reason in time_off_reasons:
            reason_label = LEAVE_REASON_LABELS.get(leave_request.leave_reason, "")

            # Calcular horas reales si start_time y end_time están definidas
            if leave_request.start_time and leave_request.end_time:
                start_dt = datetime.combine(date, leave_request.start_time)
                end_dt = datetime.combine(date, leave_request.end_time)
                duration = end_dt - start_dt
                hours = round(duration.total_seconds() / 3600, 2)
                return (reason_label, hours)
            else:
                # Si no hay horas específicas, es un día completo
                return (reason_label, work_hours_per_day)

        return None

    except Exception:
        return None


def get_holiday_hours_for_day(company_id, date):
    """Obtener horas de festivo para un día específico (si existe en CompanySettings)"""
    try:
        company_settings = CompanySettings.objects.filter(company_id=company_id).first()

        if company_settings and company_settings.holidays:
            # holidays es ArrayField de DateField
            if date in company_settings.holidays:
                return get_work_hours_per_day(company_id)

        return 0.0

    except Exception:
        return 0.0


def get_work_hours_per_day(company_id):
    """Obtener jornada laboral de la empresa (default 8h)"""
    try:
        company_settings = CompanySettings.objects.filter(company_id=company_id).first()

        if company_settings:
            # Calcular diferencia entre work_end y work_start
            start = datetime.combine(datetime.today(), company_settings.work_start)
            end = datetime.combine(datetime.today(), company_settings.work_end)
            duration = end - start
            hours = duration.total_seconds() / 3600
            return max(hours, 8.0)  # Mínimo 8 horas

        return 8.0

    except Exception:
        return 8.0


def calculate_ordinary_hours(work_time, pauses, vacations, holidays, work_hours_per_day=8):
    """Calcular horas ordinarias

    work_time: horas trabajadas (entrada - salida)
    pauses: horas de pausa
    vacations: horas de vacaciones
    holidays: horas festivos

    Fórmula: work_time - pauses - vacations - holidays
    Si > work_hours_per_day: se recorta a work_hours_per_day
    """
    if not work_time or work_time <= 0:
        return 0.0

    ordinarias = max(0, work_time - pauses - vacations - holidays)
    ordinarias = min(ordinarias, work_hours_per_day)  # Máximo: jornada laboral

    return round(ordinarias, 2)


def calculate_extra_hours(work_time, ordinary_hours, work_hours_per_day=8):
    """Calcular horas extras

    extras = work_time - ordinary_hours
    Si <= 0: = 0
    """
    if not work_time or work_time <= 0:
        return 0.0

    extras = work_time - ordinary_hours
    return round(max(0, extras), 2)


# --- Clase ExcelReportGenerator ---

class ExcelReportGenerator:

    def __init__(self, user, company, report_date):
        """
        Args:
            user: Users instance
            company: Companies instance
            report_date: date object for first day of month (2026-05-01)
        """
        self.user = user
        self.company = company
        self.report_date = report_date
        self.start_date = report_date
        # Último día del mes
        next_month = report_date + timedelta(days=31)
        self.end_date = next_month.replace(day=1) - timedelta(days=1)

        self.workbook = None
        self.worksheet = None

        # Caché de datos
        self._time_entries_cache = {}
        self._leave_requests_cache = None
        self._company_settings = None

    def generate(self) -> BytesIO:
        """Generate Excel file and return as BytesIO"""
        self._create_workbook()
        self._create_headers()
        self._populate_rows()
        self._apply_styles()
        return self._save_to_bytesio()

    def _create_workbook(self):
        """Inicializar workbook con openpyxl"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = f"{self.report_date.strftime('%B %Y')}"

    def _create_headers(self):
        """Crear fila de encabezados con estilos"""
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Ancho de columnas
            self.worksheet.column_dimensions[get_column_letter(col_idx)].width = 18

    def _get_days_in_month(self):
        """Obtener cantidad de días del mes"""
        return (self.end_date - self.start_date).days + 1

    def _get_company_settings(self):
        """Obtener settings de la empresa (cached)"""
        if self._company_settings is None:
            self._company_settings = CompanySettings.objects.filter(
                company_id=self.company.id
            ).first()
        return self._company_settings

    def get_report_data(self) -> list:
        """Obtener datos del mes como lista de dicts (para JSON/preview)"""
        days_in_month = self._get_days_in_month()
        work_hours_per_day = get_work_hours_per_day(self.company.id)
        data = []

        for day_offset in range(days_in_month):
            current_date = self.start_date + timedelta(days=day_offset)
            day_of_week_num = current_date.weekday()
            day_of_week = DAYS_OF_WEEK_ES.get(day_of_week_num, 'Día')

            time_entry = TimeEntries.objects.filter(
                user=self.user,
                company=self.company,
                date=current_date
            ).first()

            entrada = tz.localtime(time_entry.clock_in).strftime('%H:%M') if time_entry and time_entry.clock_in else ''
            salida = tz.localtime(time_entry.clock_out).strftime('%H:%M') if time_entry and time_entry.clock_out else ''

            work_time = 0
            if time_entry and time_entry.clock_in and time_entry.clock_out:
                duration = time_entry.clock_out - time_entry.clock_in
                work_time = round(duration.total_seconds() / 3600, 2)

            pauses = get_pause_hours_for_day(self.user.id, self.company.id, current_date)
            vacations = get_vacation_hours_for_day(self.user.id, self.company.id, current_date)
            time_off = get_time_off_for_day(self.user.id, self.company.id, current_date)
            sick_leave = get_sick_leave_for_day(self.user.id, self.company.id, current_date)
            holidays = get_holiday_hours_for_day(self.company.id, current_date)

            ordinarias = calculate_ordinary_hours(work_time, pauses, vacations, holidays, work_hours_per_day)
            extras = calculate_extra_hours(work_time, ordinarias, work_hours_per_day)

            # Format ausencia: "Cita médica - 8.0" o vacío
            ausencia_display = ""
            if time_off:
                reason_label, hours = time_off
                ausencia_display = f"{reason_label} - {hours}h"

            data.append({
                'fecha': current_date.strftime('%d/%m/%Y'),
                'dia_semana': day_of_week,
                'entrada': entrada,
                'salida': salida,
                'ausencia': ausencia_display,
                'vacaciones': 'X' if vacations > 0 else '',
                'baja': sick_leave,
                'festivo': 'X' if holidays > 0 else '',
                'ordinarias': ordinarias if ordinarias > 0 else '',
                'extras': extras if extras > 0 else '',
                'firma': '',
                'es_fin_semana': day_of_week_num in [5, 6],
            })

        return data

    def _populate_rows(self):
        """Llenar filas de datos para cada día del mes"""
        days_in_month = self._get_days_in_month()
        work_hours_per_day = get_work_hours_per_day(self.company.id)

        for day_offset in range(days_in_month):
            current_date = self.start_date + timedelta(days=day_offset)
            row = day_offset + 2  # Comienza en fila 2 (fila 1 = headers)

            # Columna A: Fecha
            self.worksheet.cell(row=row, column=1, value=current_date.strftime('%d/%m/%Y'))

            # Columna B: Día Semana
            day_of_week_num = current_date.weekday()
            day_of_week = DAYS_OF_WEEK_ES.get(day_of_week_num, 'Día')
            self.worksheet.cell(row=row, column=2, value=day_of_week)

            # Obtener datos para este día
            time_entry = TimeEntries.objects.filter(
                user=self.user,
                company=self.company,
                date=current_date
            ).first()

            # Columna C: Hora Entrada
            entrada = None
            salida = None
            work_time = 0

            if time_entry and time_entry.clock_in:
                entrada = tz.localtime(time_entry.clock_in).strftime('%H:%M')
                self.worksheet.cell(row=row, column=3, value=entrada)

            # Columna D: Hora Salida
            if time_entry and time_entry.clock_out:
                salida = tz.localtime(time_entry.clock_out).strftime('%H:%M')
                self.worksheet.cell(row=row, column=4, value=salida)

                # Calcular tiempo trabajado en horas
                duration = time_entry.clock_out - time_entry.clock_in
                work_time = round(duration.total_seconds() / 3600, 2)

            # Calcular pausas (se restan de horas ordinarias pero NO se muestran)
            pauses = get_pause_hours_for_day(self.user.id, self.company.id, current_date)

            # Columna E: Ausencia (time-off: cita médica, deber público/legal)
            time_off = get_time_off_for_day(self.user.id, self.company.id, current_date)
            if time_off:
                reason_label, hours = time_off
                ausencia_text = f"{reason_label} - {hours}h"
                self.worksheet.cell(row=row, column=5, value=ausencia_text)

            # Columna F: Vacaciones (mostrar "X" en lugar de horas)
            vacations = get_vacation_hours_for_day(self.user.id, self.company.id, current_date)
            if vacations > 0:
                self.worksheet.cell(row=row, column=6, value='X')

            # Columna G: Baja (sick leave: enfermedad, maternidad, matrimonio, fallecimiento)
            sick_leave = get_sick_leave_for_day(self.user.id, self.company.id, current_date)
            if sick_leave:
                self.worksheet.cell(row=row, column=7, value=sick_leave)

            # Columna H: Festivo (mostrar "X" en lugar de horas)
            holidays = get_holiday_hours_for_day(self.company.id, current_date)
            if holidays > 0:
                self.worksheet.cell(row=row, column=8, value='X')

            # Columna I: Total Horas Ordinarias
            ordinarias = calculate_ordinary_hours(work_time, pauses, vacations, holidays, work_hours_per_day)
            self.worksheet.cell(row=row, column=9, value=ordinarias)

            # Columna J: Total Horas Extras
            extras = calculate_extra_hours(work_time, ordinarias, work_hours_per_day)
            self.worksheet.cell(row=row, column=10, value=extras)

            # Columna K: Firma Trabajador (vacío - usuario lo rellena)
            self.worksheet.cell(row=row, column=11, value='')

            # Aplicar formato a fines de semana
            if day_of_week_num in [5, 6]:  # Sábado=5, Domingo=6
                self._apply_weekend_format(row)

            # Renderizar en Excel usando get_report_data()
            # (Alternativa: refactorizar después si se requiere)

    def _apply_weekend_format(self, row):
        """Aplicar formato visual a fines de semana"""
        weekend_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
        weekend_font = Font(color='FF0000', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, 12):
            cell = self.worksheet.cell(row=row, column=col)
            cell.fill = weekend_fill
            cell.font = weekend_font
            cell.alignment = center_alignment

    def _apply_styles(self):
        """Aplicar estilos: bordes, alineación, etc."""
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        center_alignment = Alignment(horizontal='center', vertical='center')

        days_in_month = self._get_days_in_month()

        # Aplicar bordes y alineación a todas las celdas de datos
        for row in range(1, days_in_month + 2):
            for col in range(1, 12):
                cell = self.worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment

                # Formato de números para columnas de horas
                if col in [5, 6, 8, 9, 10]:
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.number_format = '0.00'

    def _save_to_bytesio(self) -> BytesIO:
        """Guardar workbook a BytesIO y retornar"""
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output


# --- Clase PDFReportGenerator ---

class PDFReportGenerator:

    def __init__(self, user, company, report_date):
        self.user = user
        self.company = company
        self.report_date = report_date
        self.start_date = report_date
        next_month = report_date + timedelta(days=31)
        self.end_date = next_month.replace(day=1) - timedelta(days=1)

    def generate(self) -> BytesIO:
        """Generar PDF y retornar como BytesIO"""
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )

        styles = getSampleStyleSheet()
        elements = []

        # --- Cabecera ---
        month_names = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                       'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        month_name = f"{month_names[self.report_date.month - 1]} {self.report_date.year}"

        title_style = ParagraphStyle(
            'Title', parent=styles['Normal'],
            fontSize=14, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4
        )
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'],
            fontSize=10, fontName='Helvetica', alignment=TA_CENTER, spaceAfter=2
        )

        elements.append(Paragraph(f"Informe Mensual de Horas — {month_name}", title_style))
        elements.append(Paragraph(f"Trabajador: {self.user.username} {self.user.surname}", subtitle_style))
        elements.append(Paragraph(f"Empresa: {self.company.name}", subtitle_style))
        elements.append(Spacer(1, 0.4*cm))

        # --- Tabla ---
        data = self._build_table_data()
        col_widths = [2.2*cm, 2.4*cm, 2*cm, 2*cm, 2.8*cm, 2.5*cm, 3.2*cm, 2*cm, 2.8*cm, 2.5*cm, 3*cm]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(self._build_table_style(len(data)))
        elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output

    def _build_table_data(self) -> list:
        """Construir datos de la tabla incluyendo cabecera"""
        header = [
            'Fecha', 'Día', 'Entrada', 'Salida', 'Ausencia',
            'Vacaciones', 'Baja', 'Festivo', 'Ord.', 'Extras', 'Firma'
        ]

        rows = [header]
        days_in_month = (self.end_date - self.start_date).days + 1
        work_hours_per_day = get_work_hours_per_day(self.company.id)

        for day_offset in range(days_in_month):
            current_date = self.start_date + timedelta(days=day_offset)
            day_of_week_num = current_date.weekday()

            time_entry = TimeEntries.objects.filter(
                user=self.user,
                company=self.company,
                date=current_date
            ).first()

            entrada = tz.localtime(time_entry.clock_in).strftime('%H:%M') if time_entry and time_entry.clock_in else ''
            salida = tz.localtime(time_entry.clock_out).strftime('%H:%M') if time_entry and time_entry.clock_out else ''

            work_time = 0
            if time_entry and time_entry.clock_in and time_entry.clock_out:
                duration = time_entry.clock_out - time_entry.clock_in
                work_time = round(duration.total_seconds() / 3600, 2)

            pauses    = get_pause_hours_for_day(self.user.id, self.company.id, current_date)
            vacations = get_vacation_hours_for_day(self.user.id, self.company.id, current_date)
            time_off  = get_time_off_for_day(self.user.id, self.company.id, current_date)
            sick_leave = get_sick_leave_for_day(self.user.id, self.company.id, current_date)
            holidays  = get_holiday_hours_for_day(self.company.id, current_date)
            ordinarias = calculate_ordinary_hours(work_time, pauses, vacations, holidays, work_hours_per_day)
            extras     = calculate_extra_hours(work_time, ordinarias, work_hours_per_day)

            # Format ausencia para PDF
            ausencia_text = ""
            if time_off:
                reason_label, hours = time_off
                ausencia_text = f"{reason_label} - {hours}h"

            rows.append([
                current_date.strftime('%d/%m/%Y'),
                DAYS_OF_WEEK_ES.get(day_of_week_num, ''),
                entrada,
                salida,
                ausencia_text,
                'X' if vacations > 0 else '',
                sick_leave,
                'X' if holidays > 0 else '',
                str(ordinarias) if ordinarias > 0 else '',
                str(extras) if extras > 0 else '',
                '',  # Firma usuario
            ])

        return rows

    def _build_table_style(self, num_rows: int) -> TableStyle:
        """Definir estilos de la tabla PDF"""
        style = [
            # Cabecera
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 8),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            # Bordes
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
            # Filas de datos
            ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',   (0, 1), (-1, -1), 7.5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]

        # Fines de semana en gris con texto rojo
        days_in_month = (self.end_date - self.start_date).days + 1
        for day_offset in range(days_in_month):
            current_date = self.start_date + timedelta(days=day_offset)
            if current_date.weekday() in [5, 6]:
                row_idx = day_offset + 1 
                style += [
                    ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#DCDCDC')),
                    ('TEXTCOLOR',  (0, row_idx), (-1, row_idx), colors.red),
                    ('FONTNAME',   (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                ]

        return TableStyle(style)
